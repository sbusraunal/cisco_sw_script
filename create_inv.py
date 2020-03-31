import paramiko
import sys, os, re
from openpyxl import *
import subprocess
import getpass

ip_list=[]
device_name=[]
device_model=[]
active_devices=[]
passive_devices=[]
ssh_successful_devices=[]
ssh_failed_for_authentication_devices=[]
ssh_failed_devices=[]
port = 22

#*********************************************************excel listelerini tanımlama ve düzenleme

book="test.xlsx"
inventory="inventory.xlsx"
wb = load_workbook(book, data_only=True)
ws1 = wb["sw list"]
wb.create_sheet("sw inventory")
ws2=wb["sw inventory"]
wb.create_sheet("unreachable")
ws3=wb["unreachable"]
ws1['R1']="Switch 1 - Chassis PID"
ws1['S1']="Switch 1 - Chassis SN"
ws1['T1']="Switch 2 - Chassis PID"
ws1['U1']="Switch 2 - Chassis SN"
ws1['V1']="Switch 3 - Chassis PID"
ws1['W1']="Switch 3 - Chassis SN"
ws1['X1']="Switch 4 - Chassis PID"
ws1['Y1']="Switch 4 - Chassis SN"
ws1['Z1']="Switch 1 - Power 1 PID"
ws1['AA1']="Switch 1 - Power 1 SN"
ws1['AB1']="Switch 1 - Power 2 PID"
ws1['AC1']="Switch 1 - Power 2 SN"
ws1['AD1']="Switch 2 - Power 1 PID"
ws1['AE1']="Switch 2 - Power 1 SN"
ws1['AF1']="Switch 2 - Power 2 PID"
ws1['AG1']="Switch 2 - Power 2 SN"
ws1['AH1']="Switch 3 - Power 1 PID"
ws1['AI1']="Switch 3 - Power 1 SN"
ws1['AJ1']="Switch 3 - Power 2 PID"
ws1['AK1']="Switch 3 - Power 2 SN"
ws1['AL1']="Switch 4 - Power 1 PID"
ws1['AM1']="Switch 4 - Power 1 SN"
ws1['AN1']="Switch 4 - Power 2 PID"
ws1['AO1']="Switch 4 - Power 2 SN"
ws2['A1']="Device IP"
ws2['B1']="Device Name"
ws2['C1']="Slot"
ws2['D1']="PID"
ws2['E1']="DESCR"
ws2['F1']="SN"

#********************************************************excel'den ip adreslerini alma

def create_ip_list():
	try:
		for cell1 in ws1['P']:
			#print(str(cell1.value))
			ip_list.append(str(cell1.value))
		for cell2 in ws1['Q']:
			device_name.append(str(cell2.value))
		for cell3 in ws1['K']:
			device_model.append(str(cell3.value))
		wb.close()
		return 1
	except:
		wb.close()
		return 0

#********************************************************aktif pasif cihaz listesi olusturma

def active_passive_device_list():

	for ip in range(1,len(ip_list)):
		connectivity = check_ping(ip_list[ip])		
		if(connectivity == 1):
			active_devices.append(ip_list[ip])
		else:
			passive_devices.append(ip_list[ip])
	return

#*******************************************************erisim kontrolu

def check_ping(ip):
	try:
		response = subprocess.Popen(["ping.exe", ip], stdout = subprocess.PIPE).communicate()[0]
		#print(response)
		if ("unreachable" in str(response)):
			print (ip + "  is unreachable!")
			return 0
		elif ("timed out." in str(response)):
			print (ip + "  is unreachable!")
			return 0
		elif response == 1:
			print (ip + " is unreachable!")
			return 0
		else:
			print (ip + " is reachable!")
			return 1
	except:
		print("There was a problem with the ping test")
		return 0

#*******************************************************ssh baglantisi kontrolu

def ssh_connect_status(device_ip):

	while True:
		try:
			ssh = paramiko.SSHClient()
			ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
			ssh.connect(device_ip, port, username, password, look_for_keys=False)
			print("*"*70+"Authentication verified for "+ device_ip)
			ssh_successful_devices.append(device_ip)
			return 1
		except paramiko.AuthenticationException:
			print("*"*70+"Authentication failed when connecting to "+ device_ip)
			ssh_failed_for_authentication_devices.append(device_ip)
			return 0
		except paramiko.SSHException:
			print("*"*70+"Unable to establish SSH connection: " + device_ip)
			ssh_failed_devices.append(device_ip)
			return 0
		except:
			print("*"*70+"Unable to establish SSH connection: " + device_ip)
			ssh_failed_devices.append(device_ip)
			return 0
	ssh.close()
	return

#******************************************************inventory bilgisini alıp excel'e yazma

def get_inv(device_ip, row_number,chassis_sn_counter,power_sn_counter):

	if "Nexus" in device_model[chassis_sn_counter]:
		counter = 1
		command = "show inventory all"
		device_type_number = 1
	elif "3650" in device_model[chassis_sn_counter]:
		counter = 3 #row'daki ilk dolu yerin index numarası
		command = "show inventory"
		device_type_number = 2
	elif "IE3000" in device_model[chassis_sn_counter]:
		counter = 1 #row'daki ilk dolu yerin index numarası
		command = "show inventory"
		device_type_number = 3
	else:
		counter = 1 #row'daki ilk dolu yerin index numarası
		command = "show inventory"
		device_type_number = 4	

	ssh = paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	ssh.connect(device_ip, port, username, password, look_for_keys=False)
	#print(command)
	stdin,stdout,stderr = ssh.exec_command(command)
	output = stdout.readlines()
	ssh.close()

	row = [[]] 
	for line in output: 
	    tmp = []
	    for element in line[0:-1].split(', '):
	        tmp.append(str(element))
	    row.append(tmp)
	#print(row)
	chassis_sn_add = 0
	power_sn_add = 0
	x=row_number
	a=0

	while counter < len(row):
		descr = str(row[counter][1]).replace("DESCR: ","")
		name = str(row[counter][0]).replace("NAME: ","")
		pid = str(row[counter+1][0]).replace("PID: ","")
		#************************************************************chassis seri no
		if "c36xx" in descr.lower():
			a=1		
		elif device_type_number == 1 and "chassis" in name.lower():# n7k ve n3k için kontrol edilmeli
			a=1
		elif "fan" in descr.lower():
			a=1
		elif device_type_number != 1 and "chassis" in name.lower():# n7k ve n3k için kontrol edilmeli
			chassis_name = str(row[counter+1][0]).replace("PID: ","")
			chassis_sn = str(row[counter+1][2]).replace("SN: ","")
			ws1['R'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
			ws1['S'+str(chassis_sn_counter+1)] = str(chassis_sn)
		
		elif chassis_sn_add < 5 and descr[0]=="\"" and descr[1]=="W" and descr[2]=="S" :
			chassis_name = str(row[counter+1][0]).replace("PID: ","")
			chassis_sn = str(row[counter+1][2]).replace("SN: ","")
			if "switch 1" in name.lower():
				ws1['R'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['S'+str(chassis_sn_counter+1)] = str(chassis_sn)
			elif "switch 2" in name.lower():
				ws1['T'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['U'+str(chassis_sn_counter+1)] = str(chassis_sn)
			elif "switch 3" in name.lower():
				ws1['V'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['W'+str(chassis_sn_counter+1)] = str(chassis_sn)
			elif "switch 4" in name.lower():
				ws1['X'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['Y'+str(chassis_sn_counter+1)] = str(chassis_sn)
			else: 
				print("Stack device has more than 4 chassis, please check this device!!"+str(device_ip)+"!!")
			chassis_sn_add +=1

		elif pid[0]=="N" and pid[1]=="3" and pid[2]=="K" :
			chassis_name = str(row[counter+1][0]).replace("PID: ","")
			chassis_sn = str(row[counter+1][2]).replace(" SN: ","")
			ws1['R'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
			ws1['S'+str(chassis_sn_counter+1)] = str(chassis_sn)

		elif chassis_sn_add < 4 and descr[0]=="\"" and descr[1]=="I" and descr[2]=="E" :#IE3000
			chassis_name = str(row[counter+1][0]).replace("PID: ","")
			chassis_sn = str(row[counter+1][2]).replace("SN: ","")
			if "Module in slot 1" in name.lower():
				ws1['R'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['S'+str(chassis_sn_counter+1)] = str(chassis_sn)
			elif "Module in slot 2" in name.lower():
				ws1['T'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['U'+str(chassis_sn_counter+1)] = str(chassis_sn)
			elif "Module in slot 3" in name.lower():
				ws1['V'+str(chassis_sn_counter+1)] = str(chassis_name.replace("\"",""))
				ws1['W'+str(chassis_sn_counter+1)] = str(chassis_sn)
			else: 
				print("IE300 Stack device has more than 3 chassis, please check this device!!"+str(device_ip)+"!!")
			chassis_sn_add +=1
		
		#***********************************************************power seri no POWER SUPPLY NUUMBER İLE KONTROL EDİLEBİLİR

		elif "power" in name.lower():
			power_name=str(row[counter][0]).replace("NAME: ","")
			power_pid = str(row[counter+1][0]).replace("PID: ","")
			power_sn = str(row[counter+1][2]).replace("SN: ","")
			if "power supply 1" in name.lower():
				ws1['Z'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AA'+str(power_sn_counter+1)] = str(power_sn.replace(" ",""))
			elif "power supply 2" in name.lower():
				ws1['AB'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AC'+str(power_sn_counter+1)] = str(power_sn.replace(" ",""))
			elif "switch 1 - power supply a" in name.lower():
				ws1['Z'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AA'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 1 - power supply b" in name.lower():
				ws1['AB'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AC'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 2 - power supply a" in name.lower():
				ws1['AD'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AE'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 2 - power supply b" in name.lower():
				ws1['AF'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AG'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 3 - power supply a" in name.lower():
				ws1['AH'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AI'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 3 - power supply b" in name.lower():
				ws1['AJ'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AK'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 4 - power supply a" in name.lower():
				ws1['AL'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AM'+str(power_sn_counter+1)] = str(power_sn)
			elif "switch 4 - power Supply b" in name.lower():
				ws1['AN'+str(power_sn_counter+1)] = str(power_pid.replace("\"",""))
				ws1['AO'+str(power_sn_counter+1)] = str(power_sn)
			else: 
				print("Stack device has more than 8 power, please check this device!!"+str(device_ip)+"!!")
			power_sn_add += 1

		#*********************************************************diğer seri nolar

		else:
			str1 = str(row[counter][0]).replace("NAME: ","")
			ws2['C'+str(x)]=(str1.replace("\"",""))
			str2 = str(row[counter+1][0]).replace("PID: ","")
			ws2['D'+str(x)]=(str2.replace("\"",""))
			if "Nexus" in device_model[chassis_sn_counter]:
				str3 = str(row[counter][1]).replace(" DESCR: ","")
				ws2['E'+str(x)]=str(str3.replace("\"",""))
				str4 = str(row[counter+1][2]).replace(" SN: ","")
				ws2['F'+str(x)] = str(str4)
			else:
				str3 = str(row[counter][1]).replace("DESCR: ","")
				ws2['E'+str(x)]=str(str3.replace("\"",""))
				str4 = str(row[counter+1][2]).replace("SN: ","")
				ws2['F'+str(x)] = str(str4)
			x+=1

		counter+=3 # inventory bilgileri arasında boş satır yoksa 2'ye düşür!!! 
		
	return x

#*****************************************************işlem yapılamayan cihazlar
def print_failed_devices():

	try:
		if passive_devices:
			ws3['A1'] = "passive_devices:"
			for a in range(0,len(passive_devices)):
				ws3['A'+str(a+2)]= str(passive_devices[a])
			ws3['B1'] = "ssh_failed_for_authentication_devices:"
		if ssh_failed_for_authentication_devices:
			for b in range(0,len(ssh_failed_for_authentication_devices)):
				ws3['B'+str(b+2)]= str(ssh_failed_for_authentication_devices[b])
			ws3['C1'] = "ssh_failed_devices:"
		if ssh_failed_devices:
			for c in range(0,len(ssh_failed_devices)):
				ws3['C'+str(c+2)]= str(ssh_failed_devices[c])
		wb.save(inventory)
		wb.close()
		return
	except: 
		wb.close()
		return

#******************************************************main process
login_inf_counter = 0
login_inf=0

while login_inf == 0 and login_inf_counter < 3:
	username = input("SSH username: ")
	password = getpass.getpass('SSH Password: ')
	if username =="" or password =="":
		print("Please enter a valid username and password.")
		login_inf_counter+=1
	else:
		login_inf=1

row_number=2
if(login_inf == 1):
	read_excel_control = create_ip_list()
	if(read_excel_control == 1):	
		active_passive_device_list()
		if active_devices:
			for x in range(0,len(active_devices)):
				result_ssh_connect = ssh_connect_status(active_devices[x])
			for x in range(1,len(ip_list)):
				ws2['A'+str(row_number)]=str(ip_list[x])
				ws2['B'+str(row_number)]=str(device_name[x])
				for y in range(0,len(ssh_successful_devices)):
					if (str(ip_list[x]) == str(ssh_successful_devices[y])):
						row_number = get_inv(ssh_successful_devices[y],row_number,x,x)
						row_number -=1			
				row_number +=1	
				wb.save(inventory)	
			wb.save(inventory)
			wb.close()
		print_failed_devices()
		print("Inventory information saved to inventory.xlsx !!!!")
	else:
		print("Ip list is not defined.")
		sys.exit()
		
else:
	print("Input error!!")
	sys.exit()
